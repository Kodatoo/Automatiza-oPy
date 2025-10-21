import os
import time
from openpyxl import load_workbook

def editar_excel_remover_ultima_linha(nome_arquivo="BIOS.xlsx"):
    pasta_downloads = os.path.expanduser("~/Downloads")
    caminho_arquivo = os.path.join(pasta_downloads, nome_arquivo)
    novo_nome_arquivo = "CHAMADOS.xlsx"
    caminho_novo_arquivo = os.path.join(pasta_downloads, novo_nome_arquivo)

    tempo_max_espera = 30
    tempo_inicial = time.time()
    while not os.path.exists(caminho_arquivo):
        if time.time() - tempo_inicial > tempo_max_espera:
            print(f"[ERRO] Arquivo {nome_arquivo} não encontrado após {tempo_max_espera} segundos.")
            return
        time.sleep(1)

    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active

        ultima_linha = ws.max_row
        if ultima_linha > 1:
            ws.delete_rows(ultima_linha)
            wb.save(caminho_novo_arquivo)
            print(f"[INFO] Última linha removida e arquivo salvo como {novo_nome_arquivo}.")
        else:
            print("[AVISO] Nenhuma linha foi removida. O arquivo tem apenas uma linha.")

    except Exception as e:
        print(f"[ERRO] Ocorreu um erro ao editar o arquivo Excel: {e}")
